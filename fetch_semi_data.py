#!/usr/bin/env python3
"""
Semi Maker Revenue — Quarterly/Monthly Data Fetcher
=====================================================
Sources:
  A-share (19 companies) : AkShare → East Money quarterly income statement
  Silergy 6415 (Taiwan)  : MOPS monthly revenue API
  MPWR / NVTS (US)       : SEC EDGAR XBRL company facts API

Output: new sheet "QM_Data" appended to the master Excel file

Run:
    pip install akshare openpyxl requests pandas
    python fetch_semi_data.py
"""

import time
import logging
import warnings
from datetime import datetime
from pathlib import Path
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────

INPUT_XLSX  = "Semi_Maker_Revenue_202603.xlsx"
OUTPUT_XLSX = "Semi_Maker_Revenue_202603.xlsx"   # overwrite in place (with backup)
OUTPUT_SHEET = "QM_Data"

# Fetch range
YEAR_START = 2020
YEAR_END   = 2026

HEADERS_EM  = {"User-Agent": "Mozilla/5.0 (compatible; research-bot/1.0)"}
HEADERS_SEC = {"User-Agent": "SemiResearch contact@example.com"}  # EDGAR requires a real UA

# ── Company Universe ──────────────────────────────────────────────────────────

ASHARE_COMPANIES = [
    ("SG micro",     "300661", "Analog"),
    ("3-Peak",       "688536", "Analog"),
    ("Chipown",      "688508", "Analog"),
    ("Fortior",      "688279", "Analog"),
    ("Southchip",    "688484", "Analog"),
    ("Joulwatt",     "688141", "Analog"),
    ("Injoinic",     "688209", "Analog"),
    ("Novosense",    "688052", "Analog"),
    ("Silan",        "600460", "P&D"),
    ("CR Micro",     "688396", "P&D"),
    ("Yangjie",      "300373", "P&D"),
    ("Sino-Micro",   "600360", "P&D"),
    ("star power",   "603290", "P&D"),
    ("NCE",          "605111", "P&D"),
    ("JieJie Micro", "300623", "P&D"),
    ("Oriental",     "688261", "P&D"),
    ("Macmicst",     "688711", "P&D"),
    ("Sanan",        "600703", "P&D"),
    ("UNT",          "688469", "P&D"),
]

TAIWAN_COMPANIES = [
    ("Silergy", "6415", "Analog"),
]

US_COMPANIES = [
    ("MPWR", "MPWR", "0001280452", "Analog"),
    ("NVTS", "NVTS", "0001831868", "Analog"),
]

# ── Company Profiles (User Provided) ──────────────────────────────────────────
# Columns: R&D Staff, Total Employees, R&D Weight, Foundry / Fab, Primary Source Title, Latest Update
COMP_METADATA = {
    "Silergy": {
        "rd_staff": "1,350", "total_emp": "1,720", "rd_weight": "78.49%",
        "foundry": "TSMC / HHGrace / VIS", "source_title": "2023 Full Year Financial Report",
        "update_time": "2024-03-31"
    },
    "SG micro": {
        "rd_staff": "1,184", "total_emp": "1,600", "rd_weight": "74.09%",
        "foundry": "TSMC / HHGrace / SMIC / CRM", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-03-25"
    },
    "3-Peak": {
        "rd_staff": "520", "total_emp": "830", "rd_weight": "62.95%",
        "foundry": "TSMC / HHGrace / SMIC", "source_title": "2025 Half Year Financial Report",
        "update_time": "2025-08-15"
    },
    "Chipown": {
        "rd_staff": "272", "total_emp": "379", "rd_weight": "71.77%",
        "foundry": "SMIC / HHGACE / TSMC / VIS", "source_title": "2025 Half Year Financial Report",
        "update_time": "2025-08-20"
    },
    "Joulwatt": {
        "rd_staff": "776", "total_emp": "1,250", "rd_weight": "62.08%",
        "foundry": "SMIC / HHGACE / TSMC / VIS", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-03-30"
    },
    "Fortior": {
        "rd_staff": "—", "total_emp": "—", "rd_weight": "—",
        "foundry": "TSMC / SMIC", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-05"
    },
    "Southchip": {
        "rd_staff": "756", "total_emp": "1,106", "rd_weight": "68.35%",
        "foundry": "SMIC / HHGACE / TSMC", "source_title": "2025 Half Year Financial Report",
        "update_time": "2025-08-10"
    },
    "Novosense": {
        "rd_staff": "588", "total_emp": "1,228", "rd_weight": "47.90%",
        "foundry": "TSMC / HHGrace / SMIC / CRM", "source_title": "2025 Half Year Financial Report",
        "update_time": "2025-08-12"
    },
    # ── P&D / Discrete ────────────────────────────────────────────────────────
    "Silan": {
        "rd_staff": "2,850", "total_emp": "8,100", "rd_weight": "35.18%",
        "foundry": "IDM (6/8/12-inch fabs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-10"
    },
    "CR Micro": {
        "rd_staff": "1,200", "total_emp": "5,800", "rd_weight": "20.69%",
        "foundry": "IDM (8/12-inch fabs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-30"
    },
    "Yangjie": {
        "rd_staff": "320", "total_emp": "3,500", "rd_weight": "9.14%",
        "foundry": "IDM (4/6-inch fabs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-25"
    },
    "Sino-Micro": {
        "rd_staff": "450", "total_emp": "2,800", "rd_weight": "16.07%",
        "foundry": "IDM (4/6-inch fabs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-28"
    },
    "star power": {
        "rd_staff": "980", "total_emp": "2,600", "rd_weight": "37.69%",
        "foundry": "IDM + TSMC / SMIC (SiC)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-20"
    },
    "NCE": {
        "rd_staff": "260", "total_emp": "680", "rd_weight": "38.24%",
        "foundry": "SMIC / HHGrace", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-22"
    },
    "JieJie Micro": {
        "rd_staff": "180", "total_emp": "1,100", "rd_weight": "16.36%",
        "foundry": "IDM (4/6-inch fabs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-26"
    },
    "Oriental": {
        "rd_staff": "210", "total_emp": "420", "rd_weight": "50.00%",
        "foundry": "TSMC / HHGrace", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-18"
    },
    "Macmicst": {
        "rd_staff": "1,450", "total_emp": "4,200", "rd_weight": "34.52%",
        "foundry": "IDM (8/12-inch, IGBT/SiC)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-29"
    },
    "Sanan": {
        "rd_staff": "3,200", "total_emp": "18,000", "rd_weight": "17.78%",
        "foundry": "IDM (2/4/6-inch, LED/SiC/GaAs)", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-15"
    },
    "Injoinic": {
        "rd_staff": "185", "total_emp": "310", "rd_weight": "59.68%",
        "foundry": "TSMC / SMIC / HHGrace", "source_title": "2024 Full Year Financial Report",
        "update_time": "2025-04-20"
    },
    # ── US Companies ──────────────────────────────────────────────────────────
    "MPWR": {
        "rd_staff": "1,650", "total_emp": "2,820", "rd_weight": "58.51%",
        "foundry": "TSMC (primary)", "source_title": "2024 Annual Report (10-K)",
        "update_time": "2025-02-20"
    },
    "NVTS": {
        "rd_staff": "195", "total_emp": "310", "rd_weight": "62.90%",
        "foundry": "TSMC (GaN-on-Si)", "source_title": "2024 Annual Report (10-K)",
        "update_time": "2025-02-28"
    },
    "UNT": {
        "rd_staff": "1,050", "total_emp": "3,200", "rd_weight": "32.8%",
        "foundry": "IDM (8-inch fab)", "source_title": "2026 Q1 Report",
        "update_time": "2026-03-15"
    }
}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _period_label(date: pd.Timestamp) -> str:
    """Convert report date to period string: 2024-03-31 → 2024Q1"""
    m = date.month
    if   m == 3:  return f"{date.year}Q1"
    elif m == 6:  return f"{date.year}Q2"
    elif m == 9:  return f"{date.year}Q3"
    elif m == 12: return f"{date.year}Q4"
    return f"{date.year}M{m:02d}"


def _ytd_to_single_quarter(df: pd.DataFrame,
                            revenue_col: str,
                            netincome_col: str) -> pd.DataFrame:
    """
    A-share quarterly reports are YTD cumulative.
    Convert to single-quarter values:
        Q1 = Q1_ytd
        Q2 = H1_ytd  - Q1_ytd
        Q3 = 9M_ytd  - H1_ytd
        Q4 = FY_ytd  - 9M_ytd
    """
    df = df.sort_values("REPORT_DATE").reset_index(drop=True)
    df["year"] = df["REPORT_DATE"].dt.year
    df["qtr"]  = df["REPORT_DATE"].dt.month.map({3:1, 6:2, 9:3, 12:4})

    rows = []
    for year, grp in df.groupby("year"):
        grp = grp.sort_values("qtr").reset_index(drop=True)
        prev_rev = 0.0
        prev_ni  = 0.0
        for _, row in grp.iterrows():
            q = row["qtr"]
            rev = float(row[revenue_col]) if pd.notna(row[revenue_col]) else float("nan")
            ni  = float(row[netincome_col]) if pd.notna(row[netincome_col]) else float("nan")
            rows.append({
                "period":        f"{year}Q{q}",
                "revenue_local": (rev - prev_rev) if pd.notna(rev) else float("nan"),
                "net_income":    (ni  - prev_ni)  if pd.notna(ni)  else float("nan"),
            })
            if pd.notna(rev): prev_rev = rev
            if pd.notna(ni):  prev_ni  = ni

    return pd.DataFrame(rows)


# ── A-share Fetcher (AkShare / East Money) ────────────────────────────────────

def fetch_ashare_quarterly(name: str, code: str) -> pd.DataFrame:
    """
    Fetch quarterly income statement for one A-share company via AkShare.
    Returns tidy DataFrame with columns:
        company, code, currency, period_type, period,
        revenue_local, net_income, source, fetched_at
    """
    try:
        import akshare as ak
        raw = ak.stock_profit_sheet_by_report_em(symbol=code)
    except Exception as e:
        log.warning(f"  [{name}] AkShare fetch failed: {e}")
        return pd.DataFrame()

    # Key columns
    rev_col = "OPERATE_INCOME"          # 营业收入
    ni_col  = "PARENT_NETPROFIT"        # 归母净利润
    date_col = "REPORT_DATE"

    missing = [c for c in [date_col, rev_col, ni_col] if c not in raw.columns]
    if missing:
        # Fallback: try TOTAL_OPERATE_INCOME
        if "TOTAL_OPERATE_INCOME" in raw.columns:
            rev_col = "TOTAL_OPERATE_INCOME"
        log.warning(f"  [{name}] Missing cols {missing}, adjusted to {rev_col}")

    raw[date_col] = pd.to_datetime(raw[date_col], errors="coerce")
    raw = raw.dropna(subset=[date_col])
    raw = raw[raw[date_col].dt.year.between(YEAR_START, YEAR_END)]

    if raw.empty:
        log.warning(f"  [{name}] No data in range {YEAR_START}-{YEAR_END}")
        return pd.DataFrame()

    # Convert to M RMB (source is already in yuan, divide by 1e6)
    raw[rev_col] = pd.to_numeric(raw[rev_col], errors="coerce") / 1e6
    raw[ni_col]  = pd.to_numeric(raw[ni_col],  errors="coerce") / 1e6

    quarterly = _ytd_to_single_quarter(raw, rev_col, ni_col)
    quarterly["company"]     = name
    quarterly["code"]        = code
    quarterly["currency"]    = "M RMB"
    quarterly["period_type"] = "Q"
    quarterly["source"]      = "AkShare/EastMoney"
    quarterly["source_url"]  = f"https://www.cninfo.com.cn/new/disclosure/stock?stockCode={code}"
    quarterly["fetched_at"]  = datetime.now().strftime("%Y-%m-%d")

    return quarterly[["company","code","currency","period_type","period",
                       "revenue_local","net_income","source","source_url","fetched_at"]]


# ── Taiwan MOPS Fetcher ────────────────────────────────────────────────────────

def fetch_taiwan_monthly(name: str, code: str) -> pd.DataFrame:
    """
    Fetch monthly revenue for a Taiwan-listed company from MOPS.
    MOPS requires: POST to ajax_t21sc03 with ROC year.
    Returns monthly revenue (net income not available at monthly granularity).
    """
    url = "https://mops.twse.com.tw/mops/web/ajax_t21sc03"
    rows = []

    for year in range(YEAR_START, YEAR_END + 1):
        roc_year = year - 1911
        payload = {
            "encodeURIComponent": "1",
            "step":        "1",
            "firstin":     "1",
            "off":         "1",
            "queryName":   "co_id",
            "inpuType":    "co_id",
            "TYPEK":       "all",
            "isnew":       "false",
            "co_id":       code,
            "year":        str(roc_year),
            "season":      "00",   # 00 = monthly report
        }
        try:
            resp = requests.post(url, data=payload, headers=HEADERS_EM, timeout=15, verify=False)
            resp.raise_for_status()
            tables = pd.read_html(resp.text, encoding="utf-8")
        except Exception as e:
            log.warning(f"  [{name}] MOPS {year} failed: {e}")
            time.sleep(1)
            continue

        # MOPS returns multiple tables; find the one with 月份 column
        data_table = None
        for t in tables:
            cols = [str(c) for c in t.columns]
            if any("月份" in c or "月" in c for c in cols):
                data_table = t
                break

        if data_table is None:
            log.warning(f"  [{name}] No monthly table found for {year}")
            continue

        # Normalise columns (MOPS HTML varies slightly by year)
        data_table.columns = [str(c).strip() for c in data_table.columns]
        # Find revenue column: 當月營收 or 当月营收
        rev_col = next(
            (c for c in data_table.columns if "當月" in c or "当月" in c), None
        )
        month_col = next(
            (c for c in data_table.columns if "月份" in c or c == "月"), None
        )
        if not rev_col or not month_col:
            log.warning(f"  [{name}] Unexpected MOPS columns: {data_table.columns.tolist()}")
            continue

        for _, row in data_table.iterrows():
            month_raw = str(row[month_col]).replace("月", "").strip()
            rev_raw   = str(row[rev_col]).replace(",", "").strip()
            try:
                month = int(month_raw)
                rev   = float(rev_raw) / 1e6   # convert to M TWD
            except ValueError:
                continue
            if not (1 <= month <= 12):
                continue
            rows.append({
                "company":      name,
                "code":         code,
                "currency":     "M TWD",
                "period_type":  "M",
                "period":       f"{year}M{month:02d}",
                "revenue_local": rev,
                "net_income":   float("nan"),   # monthly NI not published
                "source":       "MOPS",
                "source_url":   f"https://mops.twse.com.tw/mops/web/t05st10?co_id={code}",
                "fetched_at":   datetime.now().strftime("%Y-%m-%d"),
            })
        log.info(f"  [{name}] MOPS {year}: {len(rows)} months so far")
        time.sleep(0.5)   # polite rate limit

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── SEC EDGAR Fetcher ─────────────────────────────────────────────────────────

_EDGAR_BASE = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"
_REVENUE_TAGS = [
    "Revenues",
    "RevenueFromContractWithCustomerExcludingAssessedTax",
    "RevenueFromContractWithCustomerIncludingAssessedTax",
    "SalesRevenueNet",
]
_NI_TAGS = ["NetIncomeLoss", "ProfitLoss"]


def _extract_edgar_series(facts: dict, tags: list, unit: str = "USD") -> pd.DataFrame:
    """Extract quarterly/annual values from EDGAR company facts for a list of candidate tags.

    Returns columns: end, start, val, form
    - start is preserved so callers can compute period duration and avoid
      conflating 10-K annual totals with single-quarter 10-Q values.
    """
    us_gaap = facts.get("facts", {}).get("us-gaap", {})
    for tag in tags:
        node = us_gaap.get(tag, {}).get("units", {}).get(unit)
        if node:
            df = pd.DataFrame(node)
            # Keep only 10-Q (quarterly) and 10-K (annual) filings
            df = df[df["form"].isin(["10-Q", "10-K"])].copy()
            df["end"]   = pd.to_datetime(df["end"],   errors="coerce")
            df["start"] = pd.to_datetime(df.get("start"), errors="coerce") \
                          if "start" in df.columns else pd.NaT
            df = df.dropna(subset=["end"])
            # Per filing-type: keep the most recent revision for each period end
            # Deduplicate 10-Q and 10-K separately to prevent 10-K overwriting 10-Q
            q_df = df[df["form"] == "10-Q"].sort_values("filed") \
                     .drop_duplicates("end", keep="last")
            k_df = df[df["form"] == "10-K"].sort_values("filed") \
                     .drop_duplicates("end", keep="last")
            df = pd.concat([q_df, k_df], ignore_index=True) \
                   .sort_values("end").reset_index(drop=True)
            return df[["end", "start", "val", "form"]]
    return pd.DataFrame()


def fetch_edgar_quarterly(name: str, ticker: str, cik: str) -> pd.DataFrame:
    """
    Fetch quarterly Revenue and Net Income from SEC EDGAR XBRL.

    Period routing by duration (end - start):
      ≤ 95 days  → single quarter  (10-Q)  → period_type "Q", label e.g. "2024Q1"
      ≥ 340 days → full fiscal year (10-K)  → period_type "A", label e.g. "2024FY"
      anything else (6-month / 9-month YTD) → skipped to avoid double-counting

    This prevents 10-K annual totals from overwriting 10-Q single-quarter Q4 values,
    which previously caused Q4 revenue to be ~4x overstated.
    """
    url = _EDGAR_BASE.format(cik=cik)
    try:
        resp = requests.get(url, headers=HEADERS_SEC, timeout=30)
        resp.raise_for_status()
        facts = resp.json()
    except Exception as e:
        log.warning(f"  [{name}] EDGAR fetch failed: {e}")
        return pd.DataFrame()

    rev_raw = _extract_edgar_series(facts, _REVENUE_TAGS)
    ni_raw  = _extract_edgar_series(facts, _NI_TAGS)

    if rev_raw.empty:
        log.warning(f"  [{name}] No revenue tag found in EDGAR facts")
        return pd.DataFrame()

    rev_raw = rev_raw.rename(columns={"val": "revenue_local", "end": "date"})
    rev_raw["revenue_local"] /= 1e6   # → M USD

    # Build NI lookup: date → value (M USD)
    ni_lookup: dict = {}
    if not ni_raw.empty:
        for _, nr in ni_raw.iterrows():
            ni_lookup[nr["end"]] = float(nr["val"]) / 1e6

    rows = []
    for _, row in rev_raw.iterrows():
        yr = row["date"].year
        if yr not in range(YEAR_START, YEAR_END + 1):
            continue

        # Compute period duration to distinguish quarter vs annual
        duration_days: int | None = None
        if pd.notna(row["start"]):
            duration_days = (row["date"] - row["start"]).days

        if duration_days is not None:
            if duration_days <= 95:
                # ── Single quarter (10-Q) ──────────────────────────────────
                period_type = "Q"
                period      = _period_label(row["date"])
            elif duration_days >= 340:
                # ── Full fiscal year (10-K) ────────────────────────────────
                # Label as "{year}FY" to avoid collision with quarterly Q4
                period_type = "A"
                period      = f"{yr}FY"
            else:
                # 6-month or 9-month YTD segment — skip to avoid double-count
                log.debug(f"  [{name}] Skipping {duration_days}-day segment ending {row['date'].date()}")
                continue
        else:
            # No start date available: fall back to form type
            if row["form"] == "10-K":
                period_type = "A"
                period      = f"{yr}FY"
            else:
                period_type = "Q"
                period      = _period_label(row["date"])

        ni_val = ni_lookup.get(row["date"], float("nan"))

        rows.append({
            "company":       name,
            "code":          ticker,
            "currency":      "M USD",
            "period_type":   period_type,
            "period":        period,
            "revenue_local": float(row["revenue_local"]),
            "net_income":    ni_val,
            "source":        "SEC EDGAR",
            "source_url":    f"https://www.sec.gov/edgar/browse/?CIK={cik}",
            "fetched_at":    datetime.now().strftime("%Y-%m-%d"),
        })

    log.info(f"  [{name}] {sum(1 for r in rows if r['period_type']=='Q')} quarters "
             f"+ {sum(1 for r in rows if r['period_type']=='A')} annual records")
    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ── Excel Writer ──────────────────────────────────────────────────────────────

_COL_ORDER = [
    "company", "code", "currency", "period_type", "period",
    "revenue_local", "net_income", "source", "source_url", "fetched_at",
]
_COL_HEADERS = [
    "Company", "Code", "Currency", "Period Type", "Period",
    "Revenue Local (M)", "Net Income (M)", "Source", "Source URL", "Fetched At",
]
_COL_WIDTHS = [16, 10, 12, 12, 12, 20, 18, 18, 30, 14]

# Color legend for period_type
_FILL_Q = PatternFill("solid", start_color="E3F1FB")   # light blue  → quarterly
_FILL_M = PatternFill("solid", start_color="E8F5E9")   # light green → monthly
_FILL_A = PatternFill("solid", start_color="FFF9E6")   # light amber → annual


def export_to_json(df: pd.DataFrame, mappings: dict, json_path: str) -> None:
    """Transform the flat DataFrame into a hierarchical JSON for the dashboard.

    Output structure per company:
        revenue      – {period: value}  local currency, M
        net_income   – {period: value}  local currency, M
        margin       – {period: ratio}  NI / Revenue as a ratio (0–1)
                       Only populated where both rev and NI are non-null and rev ≠ 0.
                       Required by dashboard renderTrends() for the NI% overlay.
        audit        – [{yr, period, rl, ni, np, st, src_url, src_name}]
    """
    import json

    # 1. Enrich with category
    df["category"] = df["company"].map(mappings)

    # 2. Group by company
    companies = {}
    for (name, code, cat), grp in df.groupby(["company", "code", "category"]):
        rev_by_period    = {}
        ni_by_period     = {}
        margin_by_period = {}   # ← was missing; needed by renderTrends()
        audit_data       = []

        grp = grp.sort_values("period")
        for _, row in grp.iterrows():
            p   = row["period"]
            rl  = float(row["revenue_local"]) if pd.notna(row["revenue_local"]) else None
            ni  = float(row["net_income"])    if pd.notna(row["net_income"])    else None

            rev_by_period[p] = rl
            ni_by_period[p]  = ni

            # Margin: derive from the same row values — single source of truth
            if rl is not None and ni is not None and rl != 0:
                margin_by_period[p] = round(ni / rl, 8)

            # audit.np: consistent with margin (percentage form)
            np_val = round(ni / rl * 100, 4) if (rl and ni is not None) else None

            audit_data.append({
                "yr":      int(p[:4]),
                "period":  p,
                "rl":      rl,
                "ni":      ni,
                "np":      np_val,
                "st":      "reported" if "2026" in p else "actual",
                "src_url": row["source_url"],
                "src_name": row["source"],
            })

        # Metadata (profile)
        meta = COMP_METADATA.get(name, {
            "rd_staff": "—", "total_emp": "—", "rd_weight": "—",
            "foundry": "—", "source_title": "Financial Report",
            "update_time": datetime.now().strftime("%Y-%m-%d"),
        })

        companies[name] = {
            "name":       name,
            "code":       code,
            "category":   cat,
            "currency":   grp["currency"].iloc[0],
            "revenue":    rev_by_period,
            "net_income": ni_by_period,
            "margin":     margin_by_period,   # ← restored
            "audit":      audit_data,
            "profile":    meta,
        }

    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(companies, f, ensure_ascii=False, indent=2)
    log.info(f"Exported {len(companies)} companies → {json_path}")


def write_qm_sheet(df: pd.DataFrame, xlsx_path: str) -> None:
    wb = load_workbook(xlsx_path)

    # Remove old sheet if exists
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]

    ws = wb.create_sheet(OUTPUT_SHEET)

    # Header row
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E4057")
    for col_idx, (h, w) in enumerate(zip(_COL_HEADERS, _COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A2"

    # Data rows
    for row_idx, record in enumerate(df[_COL_ORDER].itertuples(index=False), start=2):
        pt = record.period_type
        fill = _FILL_Q if pt == "Q" else (_FILL_M if pt == "M" else _FILL_A)
        for col_idx, val in enumerate(record, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            # Right-align numbers
            if col_idx in (6, 7):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float) and not (val != val):   # not NaN
                    cell.number_format = '#,##0.00'

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COL_HEADERS))}1"

    # Legend (below data)
    last_row = ws.max_row + 2
    ws.cell(row=last_row,   column=1, value="Legend").font = Font(bold=True)
    ws.cell(row=last_row+1, column=1, value="Q = Quarterly").fill = _FILL_Q
    ws.cell(row=last_row+2, column=1, value="M = Monthly").fill   = _FILL_M
    ws.cell(row=last_row+3, column=1, value="A = Annual").fill    = _FILL_A

    wb.save(xlsx_path)
    log.info(f"Saved {len(df)} rows → sheet '{OUTPUT_SHEET}' in {xlsx_path}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    out_path = Path(OUTPUT_XLSX)
    if not out_path.exists():
        raise FileNotFoundError(f"Input file not found: {out_path}")

    # Backup original
    backup_path = out_path.with_name(out_path.stem + "_backup" + out_path.suffix)
    import shutil
    shutil.copy2(out_path, backup_path)
    log.info(f"Backup saved to {backup_path}")

    all_frames = []

    # ── A-share (19 companies) ─────────────────────────────────────────────
    log.info("=== Fetching A-share quarterly data via AkShare ===")
    for name, code, cat in ASHARE_COMPANIES:
        log.info(f"  Fetching {name} ({code})...")
        df = fetch_ashare_quarterly(name, code)
        if not df.empty:
            all_frames.append(df)
            log.info(f"    → {len(df)} quarters")
        time.sleep(0.4)   # polite delay

    # ── Taiwan Silergy (monthly) ───────────────────────────────────────────
    log.info("=== Fetching Taiwan monthly revenue via MOPS ===")
    for name, code, cat in TAIWAN_COMPANIES:
        log.info(f"  Fetching {name} ({code})...")
        df = fetch_taiwan_monthly(name, code)
        if not df.empty:
            all_frames.append(df)
            log.info(f"    → {len(df)} months")

    # ── US companies (SEC EDGAR) ───────────────────────────────────────────
    log.info("=== Fetching US quarterly data via SEC EDGAR ===")
    for name, ticker, cik, cat in US_COMPANIES:
        log.info(f"  Fetching {name} ({ticker})...")
        df = fetch_edgar_quarterly(name, ticker, cik)
        if not df.empty:
            all_frames.append(df)
            log.info(f"    → {len(df)} quarters")
        time.sleep(0.5)

    if not all_frames:
        log.error("No data fetched. Check network / API availability.")
        return

    combined = pd.concat(all_frames, ignore_index=True)
    combined = combined.sort_values(["company", "period"]).reset_index(drop=True)

    # Deduplicate
    combined = combined.drop_duplicates(subset=["company", "period"], keep="last")

    log.info(f"\nTotal records: {len(combined)}")
    log.info(f"Companies: {combined['company'].nunique()}")
    log.info(f"Period range: {combined['period'].min()} → {combined['period'].max()}")

    # Create category mapping for JSON export
    cat_map = {}
    for n, c, cat in ASHARE_COMPANIES: cat_map[n] = cat
    for n, c, cat in TAIWAN_COMPANIES: cat_map[n] = cat
    for n, c, cat in US_COMPANIES:     cat_map[n] = cat

    write_qm_sheet(combined, str(out_path))
    export_to_json(combined, cat_map, "data.json")
    log.info("Done.")


if __name__ == "__main__":
    main()
